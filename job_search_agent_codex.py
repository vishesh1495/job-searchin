import os
import re
import time
from dataclasses import dataclass
from typing import List, Set, Tuple, Optional
from urllib.parse import quote_plus, urljoin

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

load_dotenv()


def get_env_list(key: str) -> List[str]:
    value = os.getenv(key, "")
    return [v.strip() for v in value.split(",") if v.strip()]


LINKEDIN_BASE_URL = "https://www.linkedin.com/jobs/search/"
ROLES = get_env_list("ROLES")
LOCATIONS = get_env_list("LOCATIONS")
MAX_JOBS_PER_SEARCH = int(os.getenv("MAX_JOBS_PER_SEARCH", "30"))
MAX_PAGES_PER_SEARCH = int(os.getenv("MAX_PAGES_PER_SEARCH", "3"))
OUTPUT_XLSX = os.getenv("OUTPUT_XLSX", "linkedin_jobs.xlsx")
USER_DATA_DIR = os.getenv("USER_DATA_DIR", "./playwright_profile")
HEADLESS = os.getenv("HEADLESS", "false").lower() == "true"
WAIT_MS = int(os.getenv("WAIT_MS", "2500"))
LINKEDIN_EMAIL = os.getenv("LINKEDIN_EMAIL", "").strip()
LINKEDIN_PASSWORD = os.getenv("LINKEDIN_PASSWORD", "").strip()


@dataclass
class JobRow:
    role: str
    location: str
    title: str
    company: str
    posted: str
    job_url: str
    hiring_name: str
    hiring_link: str
    source_page: int


# ================= URL / EXCEL =================

def build_search_url(role: str, location: str, start: int = 0) -> str:
    query = f"{LINKEDIN_BASE_URL}?keywords={quote_plus(role)}&location={quote_plus(location)}"
    if start > 0:
        query += f"&start={start}"
    return query


def ensure_excel(path: str):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"
    headers = [
        "Role",
        "Location",
        "Title",
        "Company",
        "Posted",
        "Job Link",
        "Hiring Manager / HR",
        "Hiring Profile Link",
        "Source Page",
        "Applied",
        "Notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(path)
    return wb, ws


def existing_links(ws) -> Set[str]:
    links = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 6 and row[5]:
            links.add(str(row[5]).strip())
    return links


def write_jobs(jobs: List[JobRow]) -> None:
    wb, ws = ensure_excel(OUTPUT_XLSX)
    existing = existing_links(ws)

    added = 0
    for job in jobs:
        if job.job_url in existing:
            continue
        ws.append([
            job.role,
            job.location,
            job.title,
            job.company,
            job.posted,
            job.job_url,
            job.hiring_name,
            job.hiring_link,
            job.source_page,
            "",
            "",
        ])
        existing.add(job.job_url)
        added += 1

    wb.save(OUTPUT_XLSX)
    print(f"Added {added} new jobs to {OUTPUT_XLSX}")


# ================= AUTH =================

def is_logged_in(page) -> bool:
    try:
        page.goto("https://www.linkedin.com/feed/", wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(1500)
        if "login" in page.url.lower() or "checkpoint" in page.url.lower() or "challenge" in page.url.lower():
            return False
        selectors = [
            "input[placeholder*='Search']",
            "button[aria-label*='Search']",
            "a[href*='/jobs/']",
            "a[href*='/mynetwork/']",
        ]
        return any(page.locator(s).count() > 0 for s in selectors) or "feed" in page.url.lower()
    except Exception:
        return False


def login_linkedin(page) -> None:
    if is_logged_in(page):
        print("Already logged in.")
        return

    if not LINKEDIN_EMAIL or not LINKEDIN_PASSWORD:
        raise ValueError("Missing LINKEDIN_EMAIL or LINKEDIN_PASSWORD in .env")

    print("Logging into LinkedIn...")
    page.goto("https://www.linkedin.com/login", wait_until="domcontentloaded", timeout=30000)
    page.wait_for_timeout(1500)
    page.locator("#username").fill(LINKEDIN_EMAIL)
    page.locator("#password").fill(LINKEDIN_PASSWORD)
    page.locator("button[type='submit']").click()
    page.wait_for_timeout(4000)

    if any(x in page.url.lower() for x in ["checkpoint", "challenge", "login"]):
        print("Complete any verification in the browser, then press Enter here.")
        input()

    if not is_logged_in(page):
        raise RuntimeError("Login did not complete successfully.")


# ================= SCRAPING HELPERS =================

def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "")).strip()


def safe_text(locator, timeout: int = 1500) -> str:
    try:
        return clean_text(locator.first.inner_text(timeout=timeout))
    except Exception:
        return ""


def safe_attr(locator, attr: str, timeout: int = 1500) -> str:
    try:
        return (locator.first.get_attribute(attr, timeout=timeout) or "").strip()
    except Exception:
        return ""


def detect_jobs_list(page):
    candidates = [
        "ul.scaffold-layout__list-container",
        "ul.jobs-search__results-list",
        "div.jobs-search-results-list",
        "div.scaffold-layout__list",
    ]
    for selector in candidates:
        try:
            loc = page.locator(selector).first
            loc.wait_for(timeout=6000)
            return loc
        except Exception:
            continue
    return None


def get_job_cards(page):
    selectors = [
        "li:has(a.job-card-list__title)",
        "li:has(a.job-card-container__link)",
        "div.job-card-container",
    ]
    for selector in selectors:
        loc = page.locator(selector)
        try:
            if loc.count() > 0:
                return loc
        except Exception:
            continue
    return None


def extract_posted(page) -> str:
    selectors = [
        "span.jobs-unified-top-card__posted-date",
        "div.job-details-jobs-unified-top-card__tertiary-description-container span",
        "span.posted-time-ago__text",
    ]
    for selector in selectors:
        text = safe_text(page.locator(selector))
        if text:
            return text

    info_blocks = [
        "div.job-details-jobs-unified-top-card__primary-description-container",
        "div.job-details-jobs-unified-top-card__tertiary-description-container",
        "div.topcard__flavor-row",
    ]
    for selector in info_blocks:
        text = safe_text(page.locator(selector))
        if text and "·" in text:
            parts = [p.strip() for p in text.split("·")]
            if len(parts) > 1:
                return parts[-1]
    return ""


def extract_hiring_contact(page) -> Tuple[str, str]:
    """
    Best-effort only.
    LinkedIn does not consistently expose a hiring manager / recruiter on every job.
    This checks common visible patterns on the job details pane.
    """
    contact_patterns = [
        "a[href*='/in/']",
        "a[href*='/recruiter/']",
    ]

    likely_container_selectors = [
        "div.jobs-poster__container",
        "div.jobs-poster",
        "section:has-text('Meet the hiring team')",
        "section:has-text('Job poster')",
        "div:has-text('Meet the hiring team')",
        "div:has-text('Job poster')",
    ]

    # First, try likely sections.
    for container_selector in likely_container_selectors:
        try:
            container = page.locator(container_selector).first
            if container.count() == 0:
                continue
            for link_selector in contact_patterns:
                anchor = container.locator(link_selector).first
                name = safe_text(anchor)
                href = safe_attr(anchor, "href")
                if name and href:
                    return name, href
        except Exception:
            continue

    # Fallback: look for a likely visible profile link in the right panel / details.
    try:
        anchors = page.locator("a[href*='/in/']")
        limit = min(anchors.count(), 20)
        for i in range(limit):
            a = anchors.nth(i)
            name = clean_text(a.inner_text(timeout=1000)) if a else ""
            href = (a.get_attribute("href", timeout=1000) or "").strip()
            if name and href and len(name.split()) <= 5:
                return name, href
    except Exception:
        pass

    return "", ""


def click_next_page(page, current_page_number: int) -> bool:
    """
    Best-effort pagination.
    First tries numbered page button, then a generic next button.
    """
    next_page_number = current_page_number + 1

    possible_next_selectors = [
        f"button[aria-label='Page {next_page_number}']",
        f"button[aria-label='Page {next_page_number} of results']",
        f"button:has-text('{next_page_number}')",
        "button[aria-label='View next page']",
        "button[aria-label='Next']",
    ]

    for selector in possible_next_selectors:
        try:
            btn = page.locator(selector).first
            if btn.count() == 0:
                continue
            btn.scroll_into_view_if_needed(timeout=2000)
            btn.click(timeout=3000)
            page.wait_for_timeout(WAIT_MS)
            return True
        except Exception:
            continue

    return False


def extract_job_from_card(page, card, role: str, location: str, source_page: int) -> Optional[JobRow]:
    try:
        card.scroll_into_view_if_needed(timeout=3000)
        page.wait_for_timeout(300)

        link_candidates = [
            "a.job-card-list__title",
            "a.job-card-container__link",
            "a[href*='/jobs/view/']",
        ]

        title = ""
        job_url = ""
        for selector in link_candidates:
            anchor = card.locator(selector).first
            title = safe_text(anchor)
            job_url = safe_attr(anchor, "href")
            if title or job_url:
                break

        if not title:
            title = safe_text(card.locator("a").first)
        if not job_url:
            job_url = safe_attr(card.locator("a[href*='/jobs/view/']").first, "href")

        company = ""
        for selector in [
            ".job-card-container__company-name",
            ".artdeco-entity-lockup__subtitle",
            ".job-card-container__primary-description",
        ]:
            company = safe_text(card.locator(selector).first)
            if company:
                break

        # Click to load details on the right panel.
        try:
            card.click(timeout=2500)
        except Exception:
            try:
                card.locator("a").first.click(timeout=2500)
            except Exception:
                pass

        page.wait_for_timeout(WAIT_MS)

        # Some cards may not expose direct href until panel loads.
        if not job_url:
            job_url = page.url if "/jobs/view/" in page.url else ""

        posted = extract_posted(page)
        hiring_name, hiring_link = extract_hiring_contact(page)

        if job_url and job_url.startswith("/"):
            job_url = urljoin("https://www.linkedin.com", job_url)
        if hiring_link and hiring_link.startswith("/"):
            hiring_link = urljoin("https://www.linkedin.com", hiring_link)

        if not title and not company and not job_url:
            return None

        return JobRow(
            role=role,
            location=location,
            title=title,
            company=company,
            posted=posted,
            job_url=job_url,
            hiring_name=hiring_name,
            hiring_link=hiring_link,
            source_page=source_page,
        )
    except Exception:
        return None


# ================= SCRAPER =================

def scrape(role: str, location: str) -> List[JobRow]:
    results: List[JobRow] = []

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            USER_DATA_DIR,
            headless=HEADLESS,
            viewport={"width": 1440, "height": 1000},
        )
        page = context.new_page()
        login_linkedin(page)

        current_page = 1
        start_offset = 0

        while current_page <= MAX_PAGES_PER_SEARCH and len(results) < MAX_JOBS_PER_SEARCH:
            search_url = build_search_url(role, location, start=start_offset)
            print(f"\nSearching: {role} | {location} | Page {current_page}")
            page.goto(search_url, wait_until="domcontentloaded", timeout=45000)
            page.wait_for_timeout(WAIT_MS)

            jobs_list = detect_jobs_list(page)
            if jobs_list is None:
                print("Could not find jobs list on this page.")
                break

            # Gentle scroll to load cards on the current page.
            for _ in range(3):
                try:
                    jobs_list.evaluate("(el) => { el.scrollTop = el.scrollHeight; }")
                except Exception:
                    page.mouse.wheel(0, 2200)
                page.wait_for_timeout(700)

            cards = get_job_cards(page)
            if cards is None:
                print("No job cards found on this page.")
                break

            page_count = min(cards.count(), MAX_JOBS_PER_SEARCH - len(results))
            print(f"Processing up to {page_count} jobs from page {current_page}...")

            for i in range(page_count):
                row = extract_job_from_card(page, cards.nth(i), role, location, current_page)
                if row and row.job_url:
                    results.append(row)
                    print(f"[{len(results)}] {row.title} | {row.company} | {row.hiring_name or 'No visible contact'}")

                if len(results) >= MAX_JOBS_PER_SEARCH:
                    break

            if len(results) >= MAX_JOBS_PER_SEARCH:
                break

            # Prefer real pagination click. If that fails, fall back to start offset.
            moved = click_next_page(page, current_page)
            current_page += 1
            start_offset += 25

            if not moved:
                # Not fatal. We'll continue using the next start offset URL.
                page.wait_for_timeout(500)

        context.close()

    return results


# ================= MAIN =================

def main() -> None:
    if not ROLES:
        raise ValueError("Add at least one role in ROLES in your .env")
    if not LOCATIONS:
        raise ValueError("Add at least one location in LOCATIONS in your .env")

    all_jobs: List[JobRow] = []

    for role in ROLES:
        for location in LOCATIONS:
            jobs = scrape(role, location)
            all_jobs.extend(jobs)

    write_jobs(all_jobs)


if __name__ == "__main__":
    main()
